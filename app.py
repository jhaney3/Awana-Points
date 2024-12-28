from flask import Flask, request, jsonify, render_template, send_file
import openpyxl
import csv
import io
import datetime
import pdfkit
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.charts.textlabels import Label
from reportlab.graphics.widgets.markers import makeMarker
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib.colors import HexColor, white, grey
from reportlab.pdfgen import canvas
from reportlab.graphics import renderPDF

app = Flask(__name__)

EXCEL_FILE = 'points.xlsx'

# Define periods and their corresponding columns
PERIODS = {
    "2024-12-04": {"prev_bal_col": 3, "add_col": 4, "sub_col": 5, "new_bal_col": 6},
    "2024-12-11": {"prev_bal_col": 6, "add_col": 7, "sub_col": 8, "new_bal_col": 9},
    "2024-12-18": {"prev_bal_col": 9, "add_col": 10, "sub_col": 11, "new_bal_col": 12},
    "2024-12-25": {"prev_bal_col": 12, "add_col": 13, "sub_col": 14, "new_bal_col": 15}
}

@app.route('/')
def index():
    return render_template('data_selection.html', periods=list(PERIODS.keys()))

@app.route('/data_selection')
def data_selection():
    # Pass the available periods to the template
    periods = list(PERIODS.keys())
    return render_template('data_selection.html', periods=periods)

@app.route('/data_entry', methods=['GET'])
def data_entry():
    # Get the selected date from query parameters
    selected_date = request.args.get('date')
    
    # Validate that the period exists
    if selected_date not in PERIODS:
        return "Invalid date selected", 400
        
    return render_template('data_entry.html', selected_date=selected_date)


@app.route('/get_data', methods=['GET'])
def get_data():
    selected_date = request.args.get('date')
    if not selected_date or selected_date not in PERIODS:
        return jsonify({"error": "Invalid date"}), 400

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    # Get the relevant columns for the selected date
    period_info = PERIODS[selected_date]
    relevant_columns = {
        1: "Name",  # Name column is always included
        period_info["prev_bal_col"]: "Previous Balance",
        period_info["add_col"]: "Add",
        period_info["sub_col"]: "Subtract",
        period_info["new_bal_col"]: "New Balance"
    }

    # Extract only the relevant headers
    headers = []
    for col_idx in sorted(relevant_columns.keys()):
        headers.append(relevant_columns[col_idx])

    data_rows = []
    # Map rows to headers
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):  # Skip empty rows
            continue

        row_dict = {}
        for col_idx, header in zip(sorted(relevant_columns.keys()), headers):
            value = row[col_idx - 1] if col_idx - 1 < len(row) else None
            row_dict[header] = value

        data_rows.append(row_dict)

    return jsonify({"headers": headers, "data": data_rows})

def capitalize_words(name):
    """Capitalize each word in the name."""
    return ' '.join(word.capitalize() for word in name.split())

@app.route('/update_period', methods=['POST'])
def update_period():
    data = request.json
    name = capitalize_words(data.get('name', '').strip())  # Capitalize and clean the name
    date_str = data.get('date')
    add_amount = float(data.get('add', 0))
    subtract_amount = float(data.get('subtract', 0))
    is_new_name = data.get('isNewName', False)
    grade = data.get('grade', '')

    if not name:
        return jsonify({"error": "Name is required"}), 400

    if date_str not in PERIODS:
        return jsonify({"error": "Date period not recognized"}), 400

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    # Find the row for the given name or create a new one
    name_col = 1
    grade_col = 2
    row_to_update = None
    
    # Check for duplicate names when adding a new name
    if is_new_name:
        for r in range(2, sheet.max_row + 1):
            existing_name = sheet.cell(row=r, column=name_col).value
            if existing_name and existing_name.strip().lower() == name.lower():
                wb.close()
                return jsonify({"error": "This name already exists in the database"}), 400

    if not is_new_name:
        # Look for existing name (case-insensitive)
        for r in range(2, sheet.max_row + 1):
            existing_name = sheet.cell(row=r, column=name_col).value
            if existing_name and existing_name.strip().lower() == name.lower():
                row_to_update = r
                # Update the name to ensure consistent capitalization
                sheet.cell(row=r, column=name_col).value = name
                break
    else:
        # Validate grade for new students
        if not grade:
            wb.close()
            return jsonify({"error": "Grade is required for new students"}), 400
            
        # Find the first empty row for new name
        for r in range(2, sheet.max_row + 2):
            if not sheet.cell(row=r, column=name_col).value:
                row_to_update = r
                sheet.cell(row=r, column=name_col).value = name
                sheet.cell(row=r, column=grade_col).value = grade
                # Initialize all balance columns to 0
                for period_info in PERIODS.values():
                    sheet.cell(row=r, column=period_info["prev_bal_col"]).value = 0
                break

    if row_to_update is None:
        return jsonify({"error": "Could not find or create row for name"}), 500

    period_info = PERIODS[date_str]
    prev_bal_col = period_info["prev_bal_col"]
    add_col = period_info["add_col"]
    sub_col = period_info["sub_col"]
    new_bal_col = period_info["new_bal_col"]

    prev_balance = sheet.cell(row=row_to_update, column=prev_bal_col).value or 0
    sheet.cell(row=row_to_update, column=add_col).value = add_amount
    sheet.cell(row=row_to_update, column=sub_col).value = subtract_amount
    new_balance = prev_balance + add_amount - subtract_amount
    sheet.cell(row=row_to_update, column=new_bal_col).value = new_balance

    # Update the next period's previous balance
    next_period_keys = list(PERIODS.keys())
    current_period_index = next_period_keys.index(date_str)
    if current_period_index < len(next_period_keys) - 1:
        next_period = next_period_keys[current_period_index + 1]
        next_period_prev_bal_col = PERIODS[next_period]["prev_bal_col"]
        sheet.cell(row=row_to_update, column=next_period_prev_bal_col).value = new_balance

    wb.save(EXCEL_FILE)
    return jsonify({"message": "Period updated successfully", "new_balance": new_balance})

@app.route('/get_attendance', methods=['GET'])
def get_attendance():
    selected_date = request.args.get('date')
    if not selected_date or selected_date not in PERIODS:
        return jsonify({"error": "Invalid or missing date"}), 400

    # Identify all earlier periods
    period_keys = list(PERIODS.keys())
    current_idx = period_keys.index(selected_date)
    
    # If there's no earlier date, return an empty data set
    if current_idx == 0:
        return render_template('attendance_report.html', 
                             selected_date=selected_date,
                             attendance_data=[])

    # Get all earlier periods
    earlier_periods = period_keys[:current_idx]

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    attendance_data = []

    for row_num in range(2, sheet.max_row + 1):
        name_value = sheet.cell(row=row_num, column=1).value
        if not name_value:
            continue  # skip blank names

        # Check if this person has any Add or Subtract in earlier periods
        attended = False
        attendance_count = 0
        for p in earlier_periods:
            p_info = PERIODS[p]
            add_col = p_info["add_col"]
            sub_col = p_info["sub_col"]
            add_val = sheet.cell(row=row_num, column=add_col).value
            sub_val = sheet.cell(row=row_num, column=sub_col).value

            # If either is non-zero or non-blank, they have "attended"
            if (add_val and add_val != 0) or (sub_val and sub_val != 0):
                attended = True
                attendance_count += 1

        attendance_data.append({
            "Name": name_value,
            "Attended": "Yes" if attended else "No",
            "Days Present": attendance_count
        })

    # Sort by name
    attendance_data.sort(key=lambda x: x["Name"])

    return render_template('attendance_report.html', 
                         selected_date=selected_date,
                         attendance_data=attendance_data)

@app.route('/export_attendance_csv', methods=['GET'])
def export_attendance_csv():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        
        # Create a StringIO object to write CSV data
        si = io.StringIO()
        writer = csv.writer(si)
        
        # Get all periods sorted by date
        period_keys = sorted(PERIODS.keys())
        
        # Write headers with dates
        headers = ['Grade', 'Name'] + period_keys + ['Total Days Present', 'Attendance %']
        writer.writerow(headers)
        
        # Create a dictionary to store attendance data by name
        attendance_by_name = {}
        
        # For each student
        for row_num in range(2, sheet.max_row + 1):
            name_value = sheet.cell(row=row_num, column=1).value
            grade_value = sheet.cell(row=row_num, column=2).value
            if not name_value:
                continue  # skip blank names
            
            # Initialize attendance record for this student
            attendance_record = [grade_value, name_value]  # Start with grade and name
            total_days = 0
            total_periods = len(period_keys)
            
            # For each period
            for period in period_keys:
                period_info = PERIODS[period]
                add_col = period_info["add_col"]
                sub_col = period_info["sub_col"]
                
                add_val = sheet.cell(row=row_num, column=add_col).value
                sub_val = sheet.cell(row=row_num, column=sub_col).value
                
                # If either Add or Subtract has a value, they attended
                attended = "X" if ((add_val and add_val != 0) or (sub_val and sub_val != 0)) else ""
                attendance_record.append(attended)
                if attended:
                    total_days += 1
            
            # Add total days present
            attendance_record.append(str(total_days))
            
            # Calculate and add attendance percentage
            attendance_percentage = (total_days / total_periods) * 100 if total_periods > 0 else 0
            attendance_record.append(f"{attendance_percentage:.1f}%")
            
            # Store in dictionary
            attendance_by_name[name_value] = attendance_record
        
        # Write all records sorted by name
        for name in sorted(attendance_by_name.keys()):
            writer.writerow(attendance_by_name[name])
        
        # Add a summary row
        total_students = len(attendance_by_name)
        if total_students > 0:
            writer.writerow([])  # Empty row for spacing
            
            # Calculate attendance totals for each date
            summary_row = ['Summary', 'Total Present']  # Summary in grade column
            for period_idx, period in enumerate(period_keys):
                total_present = sum(1 for record in attendance_by_name.values() if record[period_idx + 2] == "X")  # +2 because of grade column
                summary_row.append(f"{total_present}/{total_students}")
            
            # Calculate total days and percentage
            total_days_all = sum(int(record[-2]) for record in attendance_by_name.values())
            avg_attendance_all = (total_days_all / (total_students * total_periods)) * 100
            summary_row.append(str(total_days_all))
            summary_row.append(f"{avg_attendance_all:.1f}%")
            
            writer.writerow(summary_row)
        
        # Create the response
        output = si.getvalue()
        si.close()
        
        # Create the response with BytesIO instead of StringIO
        mem = io.BytesIO()
        mem.write(output.encode('utf-8'))
        mem.seek(0)
        
        # Generate filename with current date
        filename = f'attendance_report_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error generating CSV: {str(e)}")  # Log the error
        return jsonify({"error": "Failed to generate attendance report"}), 500

@app.route('/export_attendance_pdf', methods=['GET'])
def export_attendance_pdf():
    """
    Generates a multi-page PDF attendance report using ReportLab and
    organizes data by grade. Each grade appears on its own page with a
    table containing student names (rows), the dates of the month (columns),
    and totals. The final page provides a summary with a bar chart
    showing attendance throughout the year.
    """
    import traceback  # For printing detailed error traces

    print("DEBUG: Entered /export_attendance_pdf route.")

    try:
        print("DEBUG: Attempting to open workbook...")
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        print(f"DEBUG: Workbook '{EXCEL_FILE}' loaded successfully. sheet.max_row={sheet.max_row}")

        # Create a BytesIO buffer for the PDF
        buffer = io.BytesIO()

        # Create the PDF document (landscape letter size)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        print("DEBUG: SimpleDocTemplate created.")

        # Define basic styles
        styles = getSampleStyleSheet()

        # Title Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=20,
            alignment=1,  # Center
            textColor=HexColor('#667eea')  # Themed color
        )

        grade_title_style = ParagraphStyle(
            'GradeHeader',
            parent=styles['Heading2'],
            fontSize=20,
            spaceAfter=18,
            alignment=1,
            textColor=HexColor('#4a5568')
        )

        summary_title_style = ParagraphStyle(
            'SummaryTitle',
            parent=styles['Heading2'],
            fontSize=22,
            spaceAfter=20,
            alignment=1,
            textColor=HexColor('#4a5568')
        )

        print("DEBUG: ParagraphStyles defined.")

        # Build up the PDF elements
        elements = []

        # 1. Cover / Main Title
        elements.append(Paragraph("Attendance Report", title_style))
        elements.append(Spacer(1, 20))
        print("DEBUG: Added main title to elements.")

        # Sort grades in specific order
        grade_order = ["Kindergarten", "1st Grade", "2nd Grade", "TNT Boys", "TNT Girls"]
        print(f"DEBUG: grade_order = {grade_order}")

        # Get all period keys sorted by date
        period_keys = sorted(PERIODS.keys())
        print(f"DEBUG: period_keys = {period_keys}")

        # Helper function to convert "YYYY-MM-DD" to something like "Dec 04"
        def date_label(date_str):
            return datetime.datetime.strptime(date_str, '%Y-%m-%d').strftime('%b %d')

        # Data structure for final summary
        # We'll store total attendance by (grade -> month -> total_present)
        # and also keep track of max total so we can scale the bar chart properly.
        attendance_summary = {}
        for g in grade_order:
            attendance_summary[g] = {}
        print("DEBUG: Initialized attendance_summary dict for each grade.")

        # 2. Loop through each grade, build a table, and start on a new page
        for grade in grade_order:
            print(f"DEBUG: Processing grade = {grade}")
            # Grade Title
            elements.append(Paragraph(f"{grade} Attendance Report", grade_title_style))
            elements.append(Spacer(1, 10))

            # Create headers for the table
            # Student Name | [Dates...] | Total Present | Attendance %
            headers = (
                ["Student Name"] +
                [date_label(d) for d in period_keys] +
                ["Total Present", "Attendance %"]
            )
            print(f"DEBUG: headers for {grade} = {headers}")

            # Gather this grade's student rows
            students_data = []
            for row_num in range(2, sheet.max_row + 1):
                name_value = sheet.cell(row=row_num, column=1).value
                grade_value = sheet.cell(row=row_num, column=2).value  # Assuming Grade is in col 2

                if not name_value or grade_value != grade:
                    continue

                record = [name_value]
                total_present = 0

                # For each date in period_keys
                for date_str in period_keys:
                    period_info = PERIODS[date_str]
                    add_col = period_info["add_col"]
                    sub_col = period_info["sub_col"]

                    add_val = sheet.cell(row=row_num, column=add_col).value
                    sub_val = sheet.cell(row=row_num, column=sub_col).value

                    present = (add_val and add_val != 0) or (sub_val and sub_val != 0)
                    checkmark = "✓" if present else ""
                    record.append(checkmark)
                    if present:
                        total_present += 1

                record.append(str(total_present))
                attendance_percentage = (total_present / len(period_keys)) * 100 if period_keys else 0
                record.append(f"{attendance_percentage:.1f}%")

                students_data.append(record)

            print(f"DEBUG: students_data for {grade} has {len(students_data)} rows.")

            # If no students for this grade
            if not students_data:
                print(f"DEBUG: No attendance records found for {grade}. Adding note and page break.")
                elements.append(Paragraph("No attendance records found for this grade.", styles['Normal']))
                elements.append(PageBreak())
                continue

            # Sort students by name
            students_data.sort(key=lambda x: x[0])  # sort on student name

            # Build table data (headers + rows)
            table_data = [headers] + students_data

            # Totals row & Averages row
            totals_row = ["Totals"]
            avg_row = ["Averages"]

            total_students = len(students_data)
            total_possible_for_grade = total_students * len(period_keys)
            print(f"DEBUG: total_students={total_students}, total_possible_for_grade={total_possible_for_grade} for {grade}.")

            # For each date column
            for col_idx in range(len(period_keys)):
                present_count = sum(
                    1 for student in students_data
                    if student[col_idx + 1] == "✓"
                )
                totals_row.append(f"{present_count}/{total_students}")
                percent = (present_count / total_students * 100) if total_students else 0
                avg_row.append(f"{percent:.1f}%")

                # For summary, break out by month
                date_obj = datetime.datetime.strptime(period_keys[col_idx], '%Y-%m-%d')
                month_key = date_obj.strftime('%Y-%m')
                if month_key not in attendance_summary[grade]:
                    attendance_summary[grade][month_key] = 0
                attendance_summary[grade][month_key] += present_count

            # Summation of all the total_present columns for each student
            overall_present_sum = sum(int(s[-2]) for s in students_data)
            overall_percentage = (overall_present_sum / total_possible_for_grade) * 100 if total_possible_for_grade else 0

            # Rightmost columns in totals_row/avg_row
            totals_row.extend([str(overall_present_sum), ""])
            avg_row.extend(["", f"{overall_percentage:.1f}%"])

            # Add a blank row for spacing
            table_data.append([""] * len(headers))
            table_data.append(totals_row)
            table_data.append(avg_row)

            grade_table = Table(table_data, repeatRows=1)
            header_color = HexColor('#667eea')
            alt_row_color = HexColor('#f7fafc')
            totals_color = HexColor('#e2e8f0')

            tbl_style = TableStyle([
                # Header row style
                ('BACKGROUND', (0, 0), (-1, 0), header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Overall grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                # Alignment
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # Name column left-aligned
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ])

            # Apply alternating row background
            for i in range(1, len(table_data) - 3):
                if i % 2 == 0:
                    tbl_style.add('BACKGROUND', (0, i), (-1, i), alt_row_color)

            # Style totals rows at the end
            tbl_style.add('BACKGROUND', (0, -2), (-1, -1), totals_color)
            tbl_style.add('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold')

            grade_table.setStyle(tbl_style)
            elements.append(grade_table)
            elements.append(Spacer(1, 20))

            # Page break after each grade
            elements.append(PageBreak())

        print("DEBUG: Finished building tables for all grades. Moving on to final summary page...")

        # 3. Final Summary Page with Bar Chart
        elements.append(Paragraph("Overall Summary & Yearly Attendance", summary_title_style))
        elements.append(Spacer(1, 10))

        # Build a mini-table summarizing monthly attendance by grade
        all_months = set()
        for g in attendance_summary:
            all_months.update(attendance_summary[g].keys())

        sorted_months = sorted(list(all_months))  # e.g., ["2024-09", "2024-10", ...]
        print(f"DEBUG: sorted_months for final summary = {sorted_months}")

        # Table header: [ "Grade", M1, M2, ... ]
        summary_headers = ["Grade"] + sorted_months
        summary_data = [summary_headers]

        # bar_data[grade] = [list of monthly sums in sorted_months order]
        bar_data = {}
        max_value = 0  # for chart scaling

        for grade in grade_order:
            row_vals = [grade]
            monthly_vals = []
            for m in sorted_months:
                val = attendance_summary[grade].get(m, 0)
                monthly_vals.append(val)
                max_value = max(max_value, val)
            # Build a table row
            for v in monthly_vals:
                row_vals.append(str(v))
            summary_data.append(row_vals)
            bar_data[grade] = monthly_vals

        # Insert the table for monthly summary
        summary_table = Table(summary_data, repeatRows=1)
        sum_tbl_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4a5568')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ])
        # Alternating rows
        for i in range(1, len(summary_data)):
            if i % 2 == 0:
                sum_tbl_style.add('BACKGROUND', (0, i), (-1, i), HexColor('#f7fafc'))

        summary_table.setStyle(sum_tbl_style)
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        print("DEBUG: Summary table added to elements.")

        # Now build a bar chart from bar_data
        print(f"DEBUG: Generating bar chart with max_value = {max_value}")

        drawing_width = 600
        drawing_height = 300
        drawing = Drawing(drawing_width, drawing_height)

        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 50
        bc.width = drawing_width - 100
        bc.height = drawing_height - 100

        bc.categoryAxis.categoryNames = sorted_months

        chart_data = []
        grade_labels = []
        for grade in grade_order:
            chart_data.append(bar_data[grade])
            grade_labels.append(grade)

        bc.data = chart_data
        bc.groupSpacing = 10
        bc.barSpacing = 3

        from reportlab.lib.colors import PCMYKColor
        bar_colors = [
            PCMYKColor(100, 60, 0, 0),
            PCMYKColor(0, 60, 100, 0),
            PCMYKColor(0, 0, 0, 50),
            PCMYKColor(60, 0, 100, 0),
            PCMYKColor(0, 70, 70, 0),
        ]
        bc.bars.strokeWidth = 0.5

        for i, _ in enumerate(grade_labels):
            bc.bars[i].fillColor = bar_colors[i % len(bar_colors)]

        bc.valueAxis.valueMin = 0
        bc.valueAxis.valueMax = max_value + 5  # small buffer on top
        step_estimate = max(1, int(max_value / 5)) if max_value > 0 else 1
        bc.valueAxis.valueStep = step_estimate
        bc.valueAxis.labelTextFormat = '%d'

        bc.categoryAxis.labels.boxAnchor = 'n'
        bc.categoryAxis.labels.angle = 45
        bc.categoryAxis.labels.dy = -15

        bc_title = String(200, drawing_height - 20, "Yearly Attendance by Grade", fontSize=14, fillColor=HexColor('#4a5568'))
        drawing.add(bc_title)
        drawing.add(bc)

        legend = Legend()
        legend.x = bc.x + bc.width + 10
        legend.y = bc.y + (bc.height / 2)
        legend.deltay = 12
        legend.colorNamePairs = [
            (bc.bars[i].fillColor, grade_labels[i]) for i in range(len(grade_labels))
        ]
        drawing.add(legend)

        print("DEBUG: Bar chart and legend created, about to create flowable.")

        class DrawingFlowable(Flowable):
            def __init__(self, draw_obj, width, height):
                Flowable.__init__(self)
                self.draw_obj = draw_obj
                self.width = width
                self.height = height

            def draw(self):
                renderPDF.draw(self.draw_obj, self.canv, 0, 0)

            def wrap(self, availWidth, availHeight):
                return (self.width, self.height)

        chart_flowable = DrawingFlowable(drawing, drawing_width, drawing_height)
        elements.append(chart_flowable)
        print("DEBUG: Bar chart flowable appended to elements.")

        # Build the PDF now
        doc.build(elements)
        print("DEBUG: doc.build() completed successfully.")

        # Retrieve PDF from buffer
        pdf_data = buffer.getvalue()
        buffer.close()

        # Prepare the response
        mem = io.BytesIO()
        mem.write(pdf_data)
        mem.seek(0)

        filename = f'attendance_report_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        print(f"DEBUG: Sending PDF as '{filename}'.")

        return send_file(
            mem,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print("DEBUG: An exception occurred in /export_attendance_pdf:")
        traceback.print_exc()  # Prints full traceback to console
        return jsonify({"error": f"Failed to generate attendance report: {str(e)}"}), 500


if __name__ == '__main__':
    app.run(debug=True)

